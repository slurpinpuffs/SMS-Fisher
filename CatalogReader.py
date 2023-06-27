import openpyxl
from openpyxl import Workbook
import os
from openpyxl.styles import numbers

version = '1.2'

print(f"Hi! Welcome to the Catalog Reader ver. {version}, an SMS Fisher tool by Amani Medcroft.")
print("Please note that this tool requires the catalog to be in a very specific format.")
print("Please check your catalog spreadsheet to make sure that column AB is titled 'SRB'.")
print("If not, you may have to delete column A if the placement off by one column or columns AA-AC if the placement is off by three columns.")
print("If you encounter any bugs, issues, or have any complaints/suggestions, just let me know!")
print()
userInput = 'y'
userInput = input("Would you like to convert a whole directory? (y/n): ")
if userInput == 'y' or userInput == 'Y':
    directory = input("Please enter the filepath for the folder of catalog spreadsheets (.xlsx files) you'd like to input: ")
    directory = directory.replace('"', '')
    isNew = input("Would you like to only import items marked 'NEW' in the catalog? (y/n): ")
    exchRate = float(input("What is the exchange rate today? (1.00 EUR = X.XX USD): "))
    priceFactor = float(input("How much should the price be multiplied by? (2023 standard = 1.20): "))

    for filename in os.scandir(directory):
            file = filename.path
            # checking if it is a file
            if os.path.isfile(file):
                splitFile = file.split('\\')
                splitFile = splitFile[-1].split('.')
                fileName = splitFile[0] + '_Fishbowl'
                folder = os.path.expanduser('~\Downloads')
                catalog = openpyxl.load_workbook(file, data_only=True)
                catSheet = catalog.active
                newWorkbook = Workbook()
                newSheet = newWorkbook.active
                maxRow = catSheet.max_row
                catRow = 1
                newRow = 2

                # Initializing spreadsheet by adding column names
                newSheet['A1'] = 'PartNumber'
                newSheet['B1'] = 'PartDescription'
                newSheet['C1'] = 'PartDetails'
                newSheet['D1'] = 'UOM'
                newSheet['E1'] = 'UPC'
                newSheet['F1'] = 'PartTypeID'
                newSheet['G1'] = 'Active'
                newSheet['H1'] = 'StdCost'
                newSheet['I1'] = 'Tracks-Lot Number'
                newSheet['J1'] = 'Tracks-Revision Level'
                newSheet['K1'] = 'Tracks-Expiration Date'
                newSheet['L1'] = 'Tracks-Serial Number'
                newSheet['M1'] = 'AssetAccount'
                newSheet['N1'] = 'COGSAccount'
                newSheet['O1'] = 'AdjustmentAccount'
                newSheet['P1'] = 'ScrapAccount'
                newSheet['Q1'] = 'VarianceAccount'
                newSheet['R1'] = 'ABCCode'
                newSheet['S1'] = 'Weight'
                newSheet['T1'] = 'WeightUOM'
                newSheet['U1'] = 'Width'
                newSheet['V1'] = 'Height'
                newSheet['W1'] = 'Len'
                newSheet['X1'] = 'SizeUOM'
                newSheet['Y1'] = 'ConsumptionRate'
                newSheet['Z1'] = 'PartURL'
                newSheet['AA1'] = 'PartRevision'
                newSheet['AB1'] = 'ProductNumber'
                newSheet['AC1'] = 'ProductDescription'
                newSheet['AD1'] = 'ProductDetails'
                newSheet['AE1'] = 'Price'
                newSheet['AF1'] = 'ProductSKU'
                newSheet['AG1'] = 'ProductUPC'
                newSheet['AH1'] = 'ProductActive'
                newSheet['AI1'] = 'ProductTaxable'
                newSheet['AJ1'] = 'ProductSOItemTypeID'
                newSheet['AK1'] = 'IncomeAccount'
                newSheet['AL1'] = 'ProductWeight'
                newSheet['AM1'] = 'ProductWeightUOM'
                newSheet['AN1'] = 'ProductWidth'
                newSheet['AO1'] = 'ProductHeight'
                newSheet['AP1'] = 'ProductLen'
                newSheet['AQ1'] = 'ProductSizeUOM'
                newSheet['AR1'] = 'Vendor'
                newSheet['AS1'] = 'DefaultVendor'
                newSheet['AT1'] = 'VendorPartNumber'
                newSheet['AU1'] = 'Cost'
                newSheet['AV1'] = 'VendorUOM'

                # Func to round number to nearest five
                def fiveRound(x, base=5):
                    return base * round(x/base)

                for num in range(0, maxRow - 1):
                    # If row is marked as NEW, then:
                    # Collects all necessary information from catalog
                    new = catSheet.cell(row=catRow, column=2).value

                    if (isNew == 'y') or (isNew == 'Y'):
                        if (new != 'NEW'):
                            catRow += 1
                            continue
                    else:
                        if (new != 'NEW') and (new != 'CF'):
                            catRow += 1
                            continue
                        
                    sku = catSheet.cell(row=catRow, column=8).value
                    upc = catSheet.cell(row=catRow, column=5).value
                    if (catSheet.cell(row=catRow, column=9).value != '') and (catSheet.cell(row=catRow, column=9).value != '--'):
                        color = catSheet.cell(row=catRow, column=9).value
                    else:
                        color = 'None'
                    if (catSheet.cell(row=catRow, column=11).value != ''):
                        size = catSheet.cell(row=catRow, column=11).value
                    else:
                        size = 'UNI'
                    if(size != 'None') and (color != 'None') and (color != "--"):
                        partNumParts = [str(sku), str(color), str(size)]
                        partNum = " - ".join(partNumParts)
                    elif(size == 'None') and (color != 'None') and (color != "--"):
                        partNumParts = [str(sku), str(color), 'UNI']
                        partNum = " - ".join(partNumParts)
                    elif(size != 'None') and ((color == 'None') or (color == "--")):
                        partNumParts = [str(sku), str(size)]
                        partNum = " - ".join(partNumParts)
                    else:
                        partNumParts = [str(sku), 'UNI']
                        partNum = " - ".join(partNumParts)
                    shortDesc = catSheet.cell(row=catRow, column=6).value
                    longDesc = catSheet.cell(row=catRow, column=7).value
                    cost = float((catSheet.cell(row=catRow, column=27).value)) * exchRate
                    cost = round(cost, 2)

                    salesPrice = float(catSheet.cell(row=catRow, column=28).value) * priceFactor
                    salesPrice = fiveRound(salesPrice)
                    salesPrice = salesPrice - 0.05
                    
                    if salesPrice == 104.95:
                        salesPrice = 99.95
                    if salesPrice == 204.95:
                        salesPrice = 199.95
                    if salesPrice == 304.95:
                        salesPrice = 299.95

                    # Outputs information into new formatted worksheet
                    newSheet.cell(row=newRow, column=1).value = partNum
                    newSheet.cell(row=newRow, column=2).value = shortDesc
                    newSheet.cell(row=newRow, column=3).value = longDesc
                    newSheet.cell(row=newRow, column=5).value = str(upc)
                    newSheet.cell(row=newRow, column=7).value = 'TRUE'
                    newSheet.cell(row=newRow, column=8).value = cost
                    newSheet.cell(row=newRow, column=28).value = partNum
                    newSheet.cell(row=newRow, column=29).value = shortDesc
                    newSheet.cell(row=newRow, column=30).value = longDesc
                    newSheet.cell(row=newRow, column=31).value = salesPrice
                    newSheet.cell(row=newRow, column=34).value = 'TRUE'
                    newSheet.cell(row=newRow, column=44).value = 'Santini Maglificio Sportivo'
                    newSheet.cell(row=newRow, column=46).value = partNum
                    newSheet.cell(row=newRow, column=47).value = cost
                    newSheet[f'E{newRow}'].number_format = '0'

                    catRow += 1
                    newRow += 1

                newWorkbook.save(filename=f'{folder}\\{fileName}.xlsx')

                print()
                print(f'Success! The catalog has been converted. The output file is at "{folder}\\{fileName}.xlsx"')
                
else:
    while(userInput != 'n' and userInput != 'N'): 
        file = input("Please enter the filepath for the catalog spreadsheet (.xlsx file) you'd like to input: ")
        file = file.replace('"', '')
        isNew = input("Would you like to only import items marked 'NEW' in the catalog? (y/n): ")
        exchRate = float(input("What is the exchange rate today? (1.00 EUR = X.XX USD): "))
        priceFactor = float(input("How much should the price be multiplied by? (2023 standard = 1.20): "))
                
        splitFile = file.split('\\')
        splitFile = splitFile[-1].split('.')
        fileName = splitFile[0] + '_Fishbowl'
        folder = os.path.expanduser('~\Downloads')
        catalog = openpyxl.load_workbook(file, data_only=True)
        catSheet = catalog.active
        newWorkbook = Workbook()
        newSheet = newWorkbook.active
        maxRow = catSheet.max_row
        catRow = 1
        newRow = 2

        # Initializing spreadsheet by adding column names
        newSheet['A1'] = 'PartNumber'
        newSheet['B1'] = 'PartDescription'
        newSheet['C1'] = 'PartDetails'
        newSheet['D1'] = 'UOM'
        newSheet['E1'] = 'UPC'
        newSheet['F1'] = 'PartTypeID'
        newSheet['G1'] = 'Active'
        newSheet['H1'] = 'StdCost'
        newSheet['I1'] = 'Tracks-Lot Number'
        newSheet['J1'] = 'Tracks-Revision Level'
        newSheet['K1'] = 'Tracks-Expiration Date'
        newSheet['L1'] = 'Tracks-Serial Number'
        newSheet['M1'] = 'AssetAccount'
        newSheet['N1'] = 'COGSAccount'
        newSheet['O1'] = 'AdjustmentAccount'
        newSheet['P1'] = 'ScrapAccount'
        newSheet['Q1'] = 'VarianceAccount'
        newSheet['R1'] = 'ABCCode'
        newSheet['S1'] = 'Weight'
        newSheet['T1'] = 'WeightUOM'
        newSheet['U1'] = 'Width'
        newSheet['V1'] = 'Height'
        newSheet['W1'] = 'Len'
        newSheet['X1'] = 'SizeUOM'
        newSheet['Y1'] = 'ConsumptionRate'
        newSheet['Z1'] = 'PartURL'
        newSheet['AA1'] = 'PartRevision'
        newSheet['AB1'] = 'ProductNumber'
        newSheet['AC1'] = 'ProductDescription'
        newSheet['AD1'] = 'ProductDetails'
        newSheet['AE1'] = 'Price'
        newSheet['AF1'] = 'ProductSKU'
        newSheet['AG1'] = 'ProductUPC'
        newSheet['AH1'] = 'ProductActive'
        newSheet['AI1'] = 'ProductTaxable'
        newSheet['AJ1'] = 'ProductSOItemTypeID'
        newSheet['AK1'] = 'IncomeAccount'
        newSheet['AL1'] = 'ProductWeight'
        newSheet['AM1'] = 'ProductWeightUOM'
        newSheet['AN1'] = 'ProductWidth'
        newSheet['AO1'] = 'ProductHeight'
        newSheet['AP1'] = 'ProductLen'
        newSheet['AQ1'] = 'ProductSizeUOM'
        newSheet['AR1'] = 'Vendor'
        newSheet['AS1'] = 'DefaultVendor'
        newSheet['AT1'] = 'VendorPartNumber'
        newSheet['AU1'] = 'Cost'
        newSheet['AV1'] = 'VendorUOM'

        # Func to round number to nearest five
        def fiveRound(x, base=5):
            return base * round(x/base)

        for num in range(0, maxRow - 1):
            # If row is marked as NEW, then:
            # Collects all necessary information from catalog
            new = catSheet.cell(row=catRow, column=2).value

            if (isNew == 'y') or (isNew == 'Y'):
                if (new != 'NEW'):
                    catRow += 1
                    continue
            else:
                if (new != 'NEW') and (new != 'CF'):
                    catRow += 1
                    continue
                
            sku = catSheet.cell(row=catRow, column=8).value
            upc = catSheet.cell(row=catRow, column=5).value
            if (catSheet.cell(row=catRow, column=9).value != '') and (catSheet.cell(row=catRow, column=9).value != '--'):
                color = catSheet.cell(row=catRow, column=9).value
            else:
                color = 'None'
            if (catSheet.cell(row=catRow, column=11).value != ''):
                size = catSheet.cell(row=catRow, column=11).value
            else:
                size = 'UNI'
            if(size != 'None') and (color != 'None') and (color != "--"):
                partNumParts = [str(sku), str(color), str(size)]
                partNum = " - ".join(partNumParts)
            elif(size == 'None') and (color != 'None') and (color != "--"):
                partNumParts = [str(sku), str(color), 'UNI']
                partNum = " - ".join(partNumParts)
            elif(size != 'None') and ((color == 'None') or (color == "--")):
                partNumParts = [str(sku), str(size)]
                partNum = " - ".join(partNumParts)
            else:
                partNumParts = [str(sku), 'UNI']
                partNum = " - ".join(partNumParts)
            shortDesc = catSheet.cell(row=catRow, column=6).value
            longDesc = catSheet.cell(row=catRow, column=7).value
            cost = float((catSheet.cell(row=catRow, column=27).value)) * exchRate
            cost = round(cost, 2)

            salesPrice = float(catSheet.cell(row=catRow, column=28).value) * priceFactor
            salesPrice = fiveRound(salesPrice)
            salesPrice = salesPrice - 0.05
            
            if salesPrice == 104.95:
                salesPrice = 99.95
            if salesPrice == 204.95:
                salesPrice = 199.95
            if salesPrice == 304.95:
                salesPrice = 299.95

            # Outputs information into new formatted worksheet
            newSheet.cell(row=newRow, column=1).value = partNum
            newSheet.cell(row=newRow, column=2).value = shortDesc
            newSheet.cell(row=newRow, column=3).value = longDesc
            newSheet.cell(row=newRow, column=5).value = str(upc)
            newSheet.cell(row=newRow, column=7).value = 'TRUE'
            newSheet.cell(row=newRow, column=8).value = cost
            newSheet.cell(row=newRow, column=28).value = partNum
            newSheet.cell(row=newRow, column=29).value = shortDesc
            newSheet.cell(row=newRow, column=30).value = longDesc
            newSheet.cell(row=newRow, column=31).value = salesPrice
            newSheet.cell(row=newRow, column=34).value = 'TRUE'
            newSheet.cell(row=newRow, column=44).value = 'Santini Maglificio Sportivo'
            newSheet.cell(row=newRow, column=46).value = partNum
            newSheet.cell(row=newRow, column=47).value = cost
            newSheet[f'E{newRow}'].number_format = '0'

            catRow += 1
            newRow += 1

        newWorkbook.save(filename=f'{folder}\\{fileName}.xlsx')

        print()
        print(f'Success! The catalog has been converted. The output file is at "{folder}\\{fileName}.xlsx"')
        userInput = input("Would you like to convert another? (y/n): ")

