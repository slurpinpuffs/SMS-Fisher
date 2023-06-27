import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font
import os

version = '1.2'

print(f"Hi! Welcome to the L1 to USD tool ver. {version}, an SMS Fisher tool by Amani Medcroft.")
print("Please note that this tool requires the catalog to be in a very specific format.")
print("If the program doesn't work correctly, you may have to delete columns AA-AC if the placement is off by three columns.")
print("If you encounter any bugs, issues, or have any complaints/suggestions, just let me know!")
print()
userInput = 'y'
while(userInput != 'n' and userInput != 'N'): 
    file = input("Please enter the filepath for the catalog spreadsheet (.xlsx file) you'd like to input: ")
    file = file.replace('"', '')
    priceFactor = float(input("What factor would you like to multiply the SRP by? (SRP * X.XX): "))

    isNew = 'n'
    splitFile = file.split('\\')
    splitFile = splitFile[-1].split('.')
    fileName = splitFile[0] + '_USD'
    folder = os.path.expanduser('~\Downloads')
    saveName = f'{folder}\\{fileName}.xlsx'
    catalog = openpyxl.load_workbook(file, data_only=True)
    catSheet = catalog.active
    maxRow = catSheet.max_row
    catRow = 3
    dollarFormat = u'"$ "#,##0.00'
    offset = 0

    if catSheet['AB2'].value != 'SRP' and catSheet['AB2'].value != 'SRB':
        if catSheet['AC2'].value == 'SRP' or catSheet['AC2'].value == 'SRB':
            offset = 1

    # Func to round number to nearest five
    def fiveRound(x, base=5):
        return base * round(x/base)

    # Initializing column names
    catSheet.cell(row=2, column=27+offset).value = 'Price'
    catSheet.cell(row=2, column=28+offset).value = 'Cost'

    for num in range(0, maxRow - 1):
        new = catSheet.cell(row=catRow, column=2+offset).value

        if (new != 'NEW') and (new != 'CF'):
            catRow += 1
            continue

        salesPrice = float(catSheet.cell(row=catRow, column=28+offset).value) * priceFactor
        salesPrice = fiveRound(salesPrice)
        salesPrice = salesPrice - 0.05
        discountSalesPrice = fiveRound(salesPrice * 0.55) - 0.05
        
        if salesPrice == 104.95:
            salesPrice = 99.95
        if salesPrice == 204.95:
            salesPrice = 199.95
        if salesPrice == 304.95:
            salesPrice = 299.95

        # Outputs information into new formatted worksheet
        catSheet.cell(row=catRow, column=27+offset).value = salesPrice
        catSheet.cell(row=catRow, column=28+offset).value = discountSalesPrice
        catSheet.cell(row=catRow, column=27+offset).number_format = dollarFormat
        catSheet.cell(row=catRow, column=28+offset).number_format = dollarFormat

        catRow += 1

    if offset == 0:
        col1 = catSheet.column_dimensions['AA']
        col2 = catSheet.column_dimensions['AB']
    elif offset == 1:
        col1 = catSheet.column_dimensions['AB']
        col2 = catSheet.column_dimensions['AC']
    col1.number_format = u'$#,##0.00'
    col2.number_format = u'$#,##0.00'

    catalog.save(filename=saveName)

    print()
    print(f'Success! The catalog has been converted. The output file is at "{saveName}"')
    userInput = input("Would you like to convert another? (y/n): ")
