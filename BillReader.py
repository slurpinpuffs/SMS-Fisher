import openpyxl
from openpyxl import Workbook
import os

print("Hi! Welcome to the Bill Reader, an SMS Fisher tool by Amani Medcroft.")
print("Please note that this tool requires the catalog to be in a very specific format.")
print("Please check your bill spreadsheet to make sure that the format is correct.")
print("If you encounter any bugs, issues, or have any complaints/suggestions, just let me know!")
print()
file = input("Please enter the filepath for the catalog spreadsheet (.xlsx file) you'd like to input: ")
file = file.replace('"', '')
exchRate = float(input("What is the exchange rate today? (1.00 EUR = X.XX USD): "))

splitFile = file.split('\\')
splitFile = splitFile[-1].split('.')
fileName = splitFile[0] + '_Fishbowl'
folder = os.path.expanduser('~\Downloads')
bill = openpyxl.load_workbook(file, data_only=True)
billSheet = bill.active
newWorkbook = Workbook()
newSheet = newWorkbook.active
maxRow = billSheet.max_row
billRow = 1
newRow = 4


# Reading address and stuff.
receiverName = billSheet.cell(row=14, column=98).value
receiverAddress = billSheet.cell(row=18, column=98).value
receiverSuite = billSheet.cell(row=19, column=98).value
receiverZipCityState = billSheet.cell(row=20, column=98).value
orderNum = billSheet.cell(row=29, column=105).value.strip()

# Initializing spreadsheet by adding column names
newSheet['A1'] = "Flag"
newSheet['B1'] = "PONum"
newSheet['C1'] = "Status"
newSheet['D1'] = "VendorName"
newSheet['E1'] = "VendorContact"
newSheet['F1'] = "RemitToName"
newSheet['G1'] = "RemitToAddress"
newSheet['H1'] = "RemitToCity"
newSheet['I1'] = "RemitToState"
newSheet['J1'] = "RemitToZip"
newSheet['K1'] = "RemitToCountry"
newSheet['L1'] = "ShipToName"
newSheet['M1'] = "DeliverToName"
newSheet['N1'] = "ShipToAddress"
newSheet['O1'] = "ShipToCity"
newSheet['P1'] = "ShipToState"
newSheet['Q1'] = "ShipToZip"
newSheet['R1'] = "ShipToCountry"
newSheet['S1'] = "CarrierName"
newSheet['T1'] = "CarrierService"
newSheet['U1'] = "VendorSONum"
newSheet['V1'] = "CustomerSONum"
newSheet['W1'] = "CreatedDate"
newSheet['X1'] = "CompletedDate"
newSheet['Y1'] = "ConfirmedDate"
newSheet['Z1'] = "FulfillmentDate"
newSheet['AA1'] = "IssuedDate"
newSheet['AB1'] = "Buyer"
newSheet['AC1'] = "ShippingTerms"
newSheet['AD1'] = "PaymentTerms"
newSheet['AE1'] = "FOB"
newSheet['AF1'] = "Note"
newSheet['AG1'] = "QuickBooksClassName"
newSheet['AH1'] = "LocationGroupName"

newSheet['A2'] = "Flag"
newSheet['B2'] = "POItemType"
newSheet['C2'] = "PartNumber"
newSheet['D2'] = "VendorPartNumber"
newSheet['E2'] = "PartQuantity"
newSheet['F2'] = "FulfilledQuantity"
newSheet['G2'] = "PickedQuantity"
newSheet['H2'] = "UOM"
newSheet['I2'] = "PartPrice"
newSheet['J2'] = "FulfillmentDate"
newSheet['K2'] = "LastFulfillmentDate"
newSheet['L2'] = "RevisionLevel"
newSheet['M2'] = "Note"
newSheet['N2'] = "QuickBooksClassName"
newSheet['O2'] = "CustomerJob"

newSheet['A3'] = "PO"
newSheet['B3'] = orderNum
newSheet['C3'] = "20"
newSheet['C3'].number_format = '0'
newSheet['D3'] = "Santini Maglifico Sportivo"
newSheet['E3'] = ""
newSheet['F3'] = "Santini Maglifico Sportivo"
newSheet['G3'] = "Via Zanica 14"
newSheet['H3'] = "Bergamo (BG)"
newSheet['I3'] = ""
newSheet['J3'] = "24126"
newSheet['K3'] = "Italy"
newSheet['L3'] = receiverName
newSheet['M3'] = receiverName
newSheet['N3'] = receiverAddress
newSheet['O3'] = "Goodyear"
newSheet['P3'] = "AZ"
newSheet['Q3'] = "85338"
newSheet['R3'] = "US"
newSheet['S3'] = "UPS"
newSheet['T3'] = "2nd Day Air"
newSheet.cell(row=newRow, column=36).value = "Euro"
newSheet.cell(row=newRow, column=37).value = exchRate



for num in range(0, maxRow - 1):

    # Checks for description value. If none, skip row.
    if billSheet.cell(row=billRow, column=33).value is None:
        billRow += 1
        continue

    # Collects all necessary information from bill
    # If an item has a code, reads the size and quantity information, saving it in a dict (size:quantity of that size).
    # Also reads color (color) and first two letters of color (shortColor).
    itemCode = billSheet.cell(row=billRow, column=4).value
    if len(itemCode) > 3:
        billSizeRow = billRow + 5
        billQuanRow = billSizeRow + 1

        # If row with sizes isn't in place, searches for the size row below row with item code.
        if billSheet.cell(row=billSizeRow, column=28).value is None:
            # tempBillRow = billRow + 1
            # while len(str(billSheet.cell(row=tempBillRow, column=28).value)) < 1:
            #     tempBillRow += 1
            #     billSizeRow = tempBillRow
            #     billQuanRow = tempBillRow + 1
            billRow += 1
            continue

        sizeQuanDict = {}

        color = str(billSheet.cell(row=billSizeRow, column=6).value)
        shortColor = color[0:2]
        costEuro = billSheet.cell(row=billRow, column=112).value
        costEuro = costEuro.replace(",",".")
        costEuro = float(costEuro) * 0.94
        costEuro = round(costEuro,2)

        for tempNum in [28, 38, 42, 53, 61, 72, 83, 90, 99, 109]:
            if billSheet.cell(row=billQuanRow, column=tempNum).value is not None:
                size = billSheet.cell(row=billSizeRow, column=tempNum).value
                quantity = billSheet.cell(row=billQuanRow, column=tempNum).value.strip()
                sizeQuanDict.update({size:quantity})

    sizes = sizeQuanDict.keys()

    itemFullNames = []
    itemQuantities = []
    for sizeKey in sizes:
        itemFullNames.append(f"{itemCode} - {shortColor} - {sizeKey}")
        itemQuantities.append(sizeQuanDict[sizeKey])

    # Outputs information into new formatted worksheet
    for item in range(0, len(itemFullNames)):
        newSheet.cell(row=newRow, column=1).value = "Item"
        newSheet.cell(row=newRow, column=2).value = "10"
        newSheet.cell(row=newRow, column=2).number_format = '0'
        newSheet.cell(row=newRow, column=3).value = itemFullNames[item]
        newSheet.cell(row=newRow, column=4).value = itemFullNames[item]
        newSheet.cell(row=newRow, column=5).value = itemQuantities[item]
        newSheet.cell(row=newRow, column=8).value = "ea"
        newSheet.cell(row=newRow, column=9).value = costEuro
        
        
        newRow +=1


    billRow += 1

newWorkbook.save(filename=f'{folder}\\{fileName}.xlsx')



